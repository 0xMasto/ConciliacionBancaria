import os
import re
import unicodedata
from pathlib import Path
from datetime import datetime

import pandas as pd
from openpyxl import load_workbook

# ---- columnas objetivo (estándar Itaú) ----
COLUMNAS_ESPERADAS = [
    "Fecha", "Concepto", "Débito", "Crédito", "Saldo", "Referencia", "Destino"
]

# ---------- utilidades ----------
def _strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")

def _norm_header(x: str) -> str:
    if x is None:
        return ""
    x = str(x).strip()
    x = re.sub(r"\s+", " ", x)
    x = _strip_accents(x).lower()
    return x

def _normalize_amount(series: pd.Series) -> pd.Series:
    s = series.astype(str).str.strip()
    neg = s.str.match(r"^\(.*\)$")
    s = s.str.replace(r"^\((.*)\)$", r"\1", regex=True)
    s = s.str.replace(".", "", regex=False).str.replace(",", ".", regex=False)
    s = s.replace({"": pd.NA, "-": pd.NA})
    out = pd.to_numeric(s, errors="coerce")
    out[neg] = -out[neg].abs()
    return out

# ---------- mapeo flexible por patrones ----------
# (coincide por "contains"/regex sobre el encabezado normalizado)
HEADER_REGEX_MAP = [
    (r"^fecha($| )", "Fecha"),
    (r"^fecha .*", "Fecha"),                 # "fecha operacion", "fecha valor", etc.
    (r"concepto|descripcion|descripción", "Concepto"),
    (r"^deb(ito|itos)?(\b| )", "Débito"),    # débito, débitos, "débito (uyu)"
    (r"^cred(ito|itos)?(\b| )", "Crédito"),  # crédito, créditos, "crédito u$s"
    (r"^saldo(\b| )", "Saldo"),
    (r"referencia|nro .*ref|nro\.? referencia", "Referencia"),
    (r"destino|observaciones|detalle", "Destino"),
]

FOOTER_HINTS = [
    "saldo anterior", "saldo actual", "total ", "total:", "cantidad de movimientos"
]

def _is_footer_row(vals) -> bool:
    text = " ".join("" if v is None else _strip_accents(str(v)).lower() for v in vals)
    return any(h in text for h in FOOTER_HINTS)

# ---------- conversión (si es .xls) con Excel COM (Windows) ----------
def _convert_xls_to_xlsx_with_excel(path_xls: str) -> str:
    """Convierte .xls -> .xlsx usando Excel COM (requiere Excel en Windows)."""
    import tempfile
    import win32com.client as win32
    import pythoncom

    if not os.path.exists(path_xls):
        raise FileNotFoundError(f"No existe el archivo: {path_xls}")

    abs_in = os.path.abspath(path_xls)
    temp_dir = Path(tempfile.gettempdir()) / "itau_xls_conv"
    temp_dir.mkdir(parents=True, exist_ok=True)
    out_name = f"{Path(abs_in).stem}_conv_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}.xlsx"
    out_path = str(temp_dir / out_name)

    pythoncom.CoInitialize()
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(abs_in, ReadOnly=True)
        wb.SaveAs(out_path, FileFormat=51)  # xlsx
        wb.Close(False)
        excel.Quit()
    finally:
        pythoncom.CoUninitialize()

    if not os.path.exists(out_path):
        raise RuntimeError("Excel no produjo el archivo .xlsx de salida.")
    return out_path

def _ensure_xlsx(path: str) -> str:
    if path.lower().endswith(".xlsx"):
        return path
    if path.lower().endswith(".xls"):
        return _convert_xls_to_xlsx_with_excel(path)
    return path

# ---------- paso A: método rápido con pandas ----------
def _try_pandas_header_detection(ruta_xlsx: str) -> pd.DataFrame | None:
    # 1) localizar fila con "fecha"
    df_raw = pd.read_excel(ruta_xlsx, header=None, engine="openpyxl")
    header_row = None
    for i, fila in df_raw.iterrows():
        if fila.astype(str).str.contains(r"(?i)\bfecha\b", na=False).any():
            header_row = i
            break
    if header_row is None:
        return None

    # 2) leer con esa fila como encabezado
    df = pd.read_excel(ruta_xlsx, header=header_row, engine="openpyxl")

    # 3) normalizar nombres y mapear por regex
    original_cols = list(df.columns)
    norm_cols = [_norm_header(c) for c in original_cols]

    mapped = {}
    used = set()
    for idx, ncol in enumerate(norm_cols):
        for pattern, std in HEADER_REGEX_MAP:
            if re.search(pattern, ncol):
                if std not in mapped:  # primera coincidencia gana
                    mapped[std] = original_cols[idx]
                    used.add(idx)
                break  # pasa a la siguiente columna

    # criterios de aceptación relajados:
    must_have = ("Fecha",)
    amount_any = any(k in mapped for k in ("Débito", "Crédito", "Saldo"))
    if all(k in mapped for k in must_have) and amount_any:
        # construir dataframe final con las columnas estándar (faltantes como NA)
        out = pd.DataFrame()
        for c in COLUMNAS_ESPERADAS:
            if c in mapped:
                out[c] = df[mapped[c]]
            else:
                out[c] = pd.NA

        # tipos
        out["Fecha"] = pd.to_datetime(out["Fecha"], errors="coerce", dayfirst=True)
        out = out.dropna(subset=["Fecha"]).reset_index(drop=True)

        for col in ["Débito", "Crédito", "Saldo"]:
            if col in out.columns:
                out[col] = _normalize_amount(out[col])

        return out[COLUMNAS_ESPERADAS]

    return None

# ---------- paso B: fusión de 2–3 filas de encabezado con openpyxl ----------
def _find_header_by_row_fusion(ws) -> tuple[int, list[str]] | tuple[None, None]:
    """
    Toma las primeras ~30 filas y crea encabezados "fusionando" hasta 3 filas.
    Si logra mapear a nuestro esquema, devuelve (row_index_base, headers_fusionados).
    """
    rows = list(ws.iter_rows(values_only=True))
    top = rows[:30] if len(rows) > 30 else rows
    # probamos combinaciones de 1, 2 y 3 filas para formar un header "apilado"
    for start in range(len(top)):
        for depth in (1, 2, 3):
            if start + depth > len(top):
                continue
            # combinar celda a celda (unir con espacio lo no nulo)
            segment = top[start:start+depth]
            max_len = max(len(r) for r in segment)
            fused = []
            for j in range(max_len):
                parts = []
                for k in range(depth):
                    val = segment[k][j] if j < len(segment[k]) else None
                    if val is not None and str(val).strip():
                        parts.append(str(val).strip())
                fused.append(" ".join(parts) if parts else "")

            # mapear fused -> estándar
            mapped = {}
            for idx, h in enumerate(fused):
                norm = _norm_header(h)
                for pattern, std in HEADER_REGEX_MAP:
                    if re.search(pattern, norm):
                        if std not in mapped:
                            mapped[std] = idx
                        break

            has_fecha = "Fecha" in mapped
            has_amount = any(c in mapped for c in ("Débito", "Crédito", "Saldo"))
            has_concepto_like = any(c in mapped for c in ("Concepto", "Referencia", "Destino"))

            if has_fecha and has_amount and has_concepto_like:
                return start, fused  # esta fila (start) funciona como "header" base

    return None, None

def _table_from_fused_header(ws, header_start: int, fused_headers: list[str]) -> pd.DataFrame:
    rows = list(ws.iter_rows(values_only=True))
    data_rows = rows[header_start + 1:]
    # construir col_map con regex
    col_map = {}
    for idx, h in enumerate(fused_headers):
        norm = _norm_header(h)
        for pattern, std in HEADER_REGEX_MAP:
            if re.search(pattern, norm):
                if std not in col_map:
                    col_map[std] = idx
                break

    registros = []
    for vals in data_rows:
        # cortar en pie de tabla
        if _is_footer_row(vals):
            break

        # fila vacía respecto a columnas mapeadas
        if all((vals[col_map[c]] is None if c in col_map and col_map[c] < len(vals) else True)
               for c in col_map.keys()):
            continue

        registro = {c: (vals[col_map[c]] if c in col_map and col_map[c] < len(vals) else None)
                    for c in COLUMNAS_ESPERADAS}
        registros.append(registro)

    if not registros:
        return pd.DataFrame(columns=COLUMNAS_ESPERADAS)

    df = pd.DataFrame(registros)

    # tipos
    if "Fecha" in df.columns:
        df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce", dayfirst=True)
        df = df.dropna(subset=["Fecha"])

    for col in ["Débito", "Crédito", "Saldo"]:
        if col in df.columns:
            df[col] = _normalize_amount(df[col])

    # asegurar orden
    for c in COLUMNAS_ESPERADAS:
        if c not in df.columns:
            df[c] = pd.NA

    return df[COLUMNAS_ESPERADAS].reset_index(drop=True)

# ---------- API de lectura ----------
def leer_movimientos_itau(path_in: str) -> pd.DataFrame:
    ruta = _ensure_xlsx(path_in)

    # Paso A: intento rápido con pandas
    try:
        df_fast = _try_pandas_header_detection(ruta)
        if df_fast is not None and not df_fast.empty:
            return df_fast.reset_index(drop=True)
    except Exception:
        pass  # seguimos al paso B

    # Paso B: openpyxl + fusión de filas de encabezado
    wb = load_workbook(ruta, read_only=True, data_only=True)
    bloques = []

    for ws in wb.worksheets:
        start, fused = _find_header_by_row_fusion(ws)
        if start is not None:
            df_blk = _table_from_fused_header(ws, start, fused)
            if not df_blk.empty:
                bloques.append(df_blk)

    if not bloques:
        raise ValueError("No se detectó ninguna fila de encabezados compatible para Itaú (probé pandas y openpyxl con fusión de filas).")

    return pd.concat(bloques, ignore_index=True).dropna(how="all").reset_index(drop=True)

# ---------- API para main.py ----------
def procesar_itau(ruta: str):
    """
    Procesa un archivo Itaú y devuelve un DataFrame.
    """
    df = leer_movimientos_itau(ruta)
    print(f"✅ Procesado archivo Itaú ({len(df)} filas).")
    print(df.head())
    return df

# ---- prueba manual ----
if __name__ == "__main__":
    ruta = r"Archivos/Estado_De_Cuenta_2769087_-_2025-10-22T171649.713.xls"
    procesar_itau(ruta)
