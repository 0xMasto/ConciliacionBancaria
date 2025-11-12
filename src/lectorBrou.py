import os
import unicodedata
import re
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
from datetime import datetime

COLUMNAS_ESPERADAS = [
    "Fecha", "Descripción", "Número de documento",
    "Asunto", "Dependencia", "Débito", "Crédito"
]

# ---- utilidades ----
def _strip_accents(s: str) -> str:
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")

def _normalize_amount(series: pd.Series) -> pd.Series:
    s = series.astype(str).str.strip()
    neg = s.str.match(r"^\(.*\)$")
    s = s.str.replace(r"^\((.*)\)$", r"\1", regex=True)
    s = s.str.replace(".", "", regex=False).str.replace(",", ".", regex=False)
    s = s.replace({"": pd.NA, "-": pd.NA})
    out = pd.to_numeric(s, errors="coerce")
    out[neg] = -out[neg].abs()
    return out

def _norm_header(x: str) -> str:
    if x is None:
        return ""
    x = str(x).strip()
    x = re.sub(r"\s+", " ", x)
    x = _strip_accents(x).lower()
    return x

# ---- Variantes comunes de encabezados ----
HEADER_MAP = {
    "fecha": "Fecha",
    "fecha valor": "Fecha",
    "descripcion": "Descripción",
    "descripción": "Descripción",
    "numero de documento": "Número de documento",
    "nro de documento": "Número de documento",
    "nº de documento": "Número de documento",
    "n° de documento": "Número de documento",
    "nro documento": "Número de documento",
    "no. de documento": "Número de documento",
    "asunto": "Asunto",
    "dependencia": "Dependencia",
    "debito": "Débito",
    "débito": "Débito",
    "debitos": "Débito",
    "débitos": "Débito",
    "credito": "Crédito",
    "crédito": "Crédito",
    "creditos": "Crédito",
    "créditos": "Crédito",
}

FOOTER_HINTS = ["saldo anterior", "saldo actual", "total ", "total:"]

def _is_footer_row(vals) -> bool:
    text = " ".join("" if v is None else _strip_accents(str(v)).lower() for v in vals)
    return any(h in text for h in FOOTER_HINTS)

# ---- conversión (si es .xls) ----
def _convert_xls_to_xlsx_with_excel(path_xls: str) -> str:
    """Convierte .xls -> .xlsx usando Excel COM (solo en Windows con Excel instalado)."""
    import tempfile
    import win32com.client as win32
    import pythoncom

    if not os.path.exists(path_xls):
        raise FileNotFoundError(f"No existe el archivo: {path_xls}")

    abs_in = os.path.abspath(path_xls)
    temp_dir = Path(tempfile.gettempdir()) / "brou_xls_conv"
    temp_dir.mkdir(parents=True, exist_ok=True)
    out_name = f"{Path(abs_in).stem}_conv_{datetime.now().strftime('%Y%m%d_%H%M%S_%f')}.xlsx"
    out_path = str(temp_dir / out_name)

    pythoncom.CoInitialize()
    try:
        excel = win32.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        wb = excel.Workbooks.Open(abs_in, ReadOnly=True)
        wb.SaveAs(out_path, FileFormat=51)  # xlsx
        wb.Close(False)
        excel.Quit()
    finally:
        pythoncom.CoUninitialize()

    if not os.path.exists(out_path):
        raise RuntimeError("Excel no produjo el archivo .xlsx de salida.")
    return out_path

def _ensure_xlsx(path: str) -> str:
    if path.lower().endswith(".xlsx"):
        return path
    if path.lower().endswith(".xls"):
        return _convert_xls_to_xlsx_with_excel(path)
    return path

# ---- detección de encabezados ----
def _find_header_row_and_colmap(rows) -> tuple[int, dict] | tuple[None, None]:
    for i, r in enumerate(rows):
        norm_cells = [_norm_header(v) for v in r]
        col_map = {}
        for j, key in enumerate(norm_cells):
            std = HEADER_MAP.get(key)
            if std and std not in col_map:
                col_map[std] = j

        if len(col_map) >= 4 and "Fecha" in col_map and (("Débito" in col_map) or ("Crédito" in col_map)):
            return i, col_map

        if len(col_map) >= 3 and "Fecha" in col_map and any(c in col_map for c in ["Descripción", "Número de documento", "Asunto"]):
            return i, col_map

    return None, None

def _build_table_from_header(rows, header_idx: int, col_map: dict) -> pd.DataFrame:
    data_rows = rows[header_idx + 1:]
    registros = []
    for vals in data_rows:
        if _is_footer_row(vals):
            break
        if all((vals[col_map[c]] is None if c in col_map else True) for c in col_map.keys()):
            continue
        registro = {col: vals[col_map[col]] if col in col_map else None for col in COLUMNAS_ESPERADAS}
        registros.append(registro)

    if not registros:
        return pd.DataFrame(columns=COLUMNAS_ESPERADAS)

    df = pd.DataFrame(registros)

    if "Fecha" in df.columns:
        df["Fecha"] = pd.to_datetime(df["Fecha"], errors="coerce", dayfirst=True)
        df = df.dropna(subset=["Fecha"])

    for col in ["Débito", "Crédito"]:
        if col in df.columns:
            df[col] = _normalize_amount(df[col])

    for c in COLUMNAS_ESPERADAS:
        if c not in df.columns:
            df[c] = pd.NA

    return df[COLUMNAS_ESPERADAS].reset_index(drop=True)

# ---- función principal ----
def leer_movimientos_brou(path_in: str) -> pd.DataFrame:
    ruta = _ensure_xlsx(path_in)
    wb = load_workbook(ruta, read_only=True, data_only=True)
    bloques = []

    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        header_idx, col_map = _find_header_row_and_colmap(rows)
        if header_idx is not None:
            df_blk = _build_table_from_header(rows, header_idx, col_map)
            if not df_blk.empty:
                bloques.append(df_blk)

    if not bloques:
        raise ValueError("No se detectó ninguna fila de encabezados compatible (Fecha, Débito/Crédito, etc.).")

    return pd.concat(bloques, ignore_index=True).dropna(how="all").reset_index(drop=True)

# ---- API compatible con main.py ----
def procesar_brou(ruta: str):
    """
    Procesa un archivo BROU y devuelve un DataFrame.
    """
    df = leer_movimientos_brou(ruta)
    print(f"✅ Procesado archivo BROU ({len(df)} filas).")
    print(df.head())
    return df

# ---- Ejemplo de uso manual ----
if __name__ == "__main__":
    ruta = r"Archivos/Detalle_Movimiento_Cuenta_-_2025-10-22T171522.294.xls"
    procesar_brou(ruta)
